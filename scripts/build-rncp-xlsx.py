from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# (bloc, code, competence, priorite, projet/preuve)
rows = [
    ("BC01 — Analyser les besoins", "BC01.1", "Reformuler la demande d'un client pour clarifier et formaliser le besoin", "Negliger pour l'instant", ""),
    ("BC01 — Analyser les besoins", "BC01.2", "Exprimer et traduire les besoins en exigences fonctionnelles et non fonctionnelles", "A developper plus tard", ""),
    ("BC01 — Analyser les besoins", "BC01.3", "Estimer les ressources necessaires (RH, competences, technologies, budget, delais)", "Negliger pour l'instant", ""),
    ("BC01 — Analyser les besoins", "BC01.4", "Estimer l'impact environnemental et societal potentiel d'une solution", "Negliger pour l'instant", ""),
    ("BC01 — Analyser les besoins", "BC01.5", "Elaborer et rediger un cahier des charges", "Negliger pour l'instant", ""),
    ("BC02 — Concevoir et modeliser", "BC02.1", "Evaluer et selectionner les solutions existantes (contraintes techniques, eco, humaines)", "Prioriser", "ADR, Model Selection Guide"),
    ("BC02 — Concevoir et modeliser", "BC02.2", "Integrer l'experience utilisateur (UX) dans toutes les dimensions du projet", "A developper plus tard", "Web TI402"),
    ("BC02 — Concevoir et modeliser", "BC02.3", "Adopter les bonnes pratiques de securite et/ou de fiabilite dans la conception", "Prioriser", "Circuit Breaker, Watchdog"),
    ("BC02 — Concevoir et modeliser", "BC02.4", "Demontrer la faisabilite technique d'un dispositif", "Prioriser", "Ecosystem (POC)"),
    ("BC02 — Concevoir et modeliser", "BC02.5", "Anticiper l'evolutivite d'une solution et son interoperabilite des la conception", "Prioriser", "Architecture Ecosystem"),
    ("BC02 — Concevoir et modeliser", "BC02.6", "Elaborer des maquettes ou des POC", "Prioriser", "Ecosystem, maquette Web"),
    ("BC03 — Mettre en oeuvre", "BC03.1", "Implementer les composants d'une architecture ou d'un systeme", "Prioriser", "Ecosystem, Chatbot, Web TI402"),
    ("BC03 — Mettre en oeuvre", "BC03.2", "Integrer les composants (compatibilite, integrite, interoperabilite, securite)", "Prioriser", "Ecosystem"),
    ("BC03 — Mettre en oeuvre", "BC03.3", "Tester et valider un dispositif logiciel, systeme ou embarque", "A developper plus tard", ""),
    ("BC03 — Mettre en oeuvre", "BC03.4", "Optimiser les performances (couts, qualite, impact env/social)", "Prioriser", "Budget Guard, Token Optimization"),
    ("BC03 — Mettre en oeuvre", "BC03.5", "Produire la documentation technique", "Prioriser", "Vault / jardin numerique"),
    ("BC03 — Mettre en oeuvre", "BC03.6", "Mettre a disposition une solution (config, install, automatisation, sauvegarde/secours)", "A developper plus tard", "Watchdog, deploiement"),
    ("BC04 — Gerer l'exploitation", "BC04.1", "Analyser les indicateurs de performance, reperer les anomalies", "A developper plus tard", "Dashboard Ecosystem"),
    ("BC04 — Gerer l'exploitation", "BC04.2", "Superviser l'utilisation pour le maintien en conditions operationnelles", "A developper plus tard", "Watchdog"),
    ("BC04 — Gerer l'exploitation", "BC04.3", "Proposer des ameliorations (fonctionnalites, performances)", "A developper plus tard", "Roadmap Ecosystem"),
    ("BC04 — Gerer l'exploitation", "BC04.4", "Mettre en oeuvre les ameliorations et mettre a jour la documentation", "A developper plus tard", ""),
    ("BC05 — OEuvrer en mode projet", "BC05.1", "Mener et partager une recherche documentaire et une veille technique", "Prioriser", ""),
    ("BC05 — OEuvrer en mode projet", "BC05.2", "Communiquer efficacement a l'ecrit et a l'oral", "Prioriser", "Portfolio bilingue, presentation Web"),
    ("BC05 — OEuvrer en mode projet", "BC05.3", "Respecter la methodologie de gestion de projet / delivery", "A developper plus tard", ""),
    ("BC05 — OEuvrer en mode projet", "BC05.4", "Realiser un suivi de l'avancement des travaux / des ressources", "A developper plus tard", "Planning, Trello"),
    ("BC05 — OEuvrer en mode projet", "BC05.5", "Collaborer en equipe pluridisciplinaire et internationale", "Prioriser", "Chatbot (180 commits), echange, langues"),
    ("BC05 — OEuvrer en mode projet", "BC05.6", "Prendre des decisions operationnelles et en comprendre les impacts", "Negliger pour l'instant", ""),
    ("BC05 — OEuvrer en mode projet", "BC05.7", "Sensibiliser aux enjeux de la TEDS (transition ecologique)", "Negliger pour l'instant", ""),
    ("BC05 — OEuvrer en mode projet", "BC05.8", "Produire la documentation pour la formation des utilisateurs", "A developper plus tard", ""),
]

wb = Workbook()
ws = wb.active
ws.title = "RNCP 41030"

headers = ["Bloc", "Code", "Competence", "Priorite", "Deja developpe", "A developper (prevu)", "Projet / preuve", "Notes"]
ws.append(headers)
hfill = PatternFill("solid", fgColor="1F2937")
for c in ws[1]:
    c.fill = hfill
    c.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    c.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
ws.row_dimensions[1].height = 30

prio = {
    "Prioriser": ("C6EFCE", "006100"),
    "A developper plus tard": ("FFEB9C", "9C6500"),
    "Negliger pour l'instant": ("EDEDED", "6B7280"),
}
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for bloc, code, comp, p, ev in rows:
    ws.append([bloc, code, comp, p, "", "", ev, ""])
    i = ws.max_row
    for col in range(1, 9):
        cell = ws.cell(row=i, column=col)
        cell.font = Font(name="Arial", size=10)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border
    pc = ws.cell(row=i, column=4)
    fg, fn = prio[p]
    pc.fill = PatternFill("solid", fgColor=fg)
    pc.font = Font(name="Arial", size=10, bold=True, color=fn)
    pc.alignment = Alignment(vertical="top", horizontal="center", wrap_text=True)
    for col in (2, 5, 6):
        ws.cell(row=i, column=col).alignment = Alignment(vertical="top", horizontal="center")

for col, w in {"A": 26, "B": 9, "C": 64, "D": 20, "E": 15, "F": 20, "G": 34, "H": 22}.items():
    ws.column_dimensions[col].width = w

dv = DataValidation(type="list", formula1='"X"', allow_blank=True)
ws.add_data_validation(dv)
dv.add(f"E2:F{ws.max_row}")

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:H{ws.max_row}"

import os
out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "RNCP-competences.xlsx")
wb.save(out)
print("saved", ws.max_row - 1, "competences ->", out)
